import httpx
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

def obtener_todas_las_tasas_bcv():
    url = "https://www.bcv.org.ve/"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    
    # Mapeo de IDs del HTML del BCV a nombres legibles y códigos ISO.
    monedas_target = {
        "dolar": {"codigo": "USD", "nombre": "Dólar estadounidense"},
        "euro": {"codigo": "EUR", "nombre": "Euro"},
        "yuan": {"codigo": "CNY", "nombre": "Yuan chino"},
        "lira": {"codigo": "TRY", "nombre": "Lira turca"},
        "rublo": {"codigo": "RUB", "nombre": "Rublo ruso"}
    }
    
    tasas_extraidas = []

    try:
        # Petición HTTP usando tu configuración
        with httpx.Client(verify=False, timeout=12.0) as client:
            respuesta = client.get(url, headers=headers)
            
            if respuesta.status_code != 200:
                print(f"Error al conectar con el BCV: Estado {respuesta.status_code}")
                return None
                
            soup = BeautifulSoup(respuesta.text, 'lxml')
            
            for id_html, info in monedas_target.items():
                bloque_moneda = soup.find(id=id_html)
                if bloque_moneda:
                    elemento_tasa = bloque_moneda.find("strong", class_="strong-tb")
                    if elemento_tasa:
                        tasa_clean = elemento_tasa.text.strip().replace('.', '').replace(',', '.')
                        tasas_extraidas.append({
                            "codigo": info["codigo"],
                            "nombre": info["nombre"],
                            "tasa": float(tasa_clean)
                        })

            return tasas_extraidas

    except Exception as e:
        print(f"Error durante el scraping: {e}")
        return None

def exportar_a_excel(tasas, nombre_archivo="Tasas_BCV_Oficiales.xlsx"):
    if not tasas:
        print("No hay datos para exportar.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Cotizaciones BCV"
    ws.views.sheetView[0].showGridLines = True

    # Banner superior
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "BANCO CENTRAL DE VENEZUELA - COTIZACIONES OFICIALES"
    title_cell.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Encabezados
    headers = ["Código ISO", "Moneda / Divisa", "Tasa Oficial (VES)", "Fecha de Extracción"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 24

    # Estilos de celda
    border_style = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    fecha_hoy = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Filas de datos
    for idx, item in enumerate(tasas):
        row_idx = idx + 4
        
        ws.cell(row=row_idx, column=1, value=item["codigo"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=item["nombre"]).alignment = Alignment(horizontal="left")
        
        # Formato numérico en moneda de Excel
        tasa_cell = ws.cell(row=row_idx, column=3, value=item["tasa"])
        tasa_cell.alignment = Alignment(horizontal="right")
        tasa_cell.number_format = '#,##0.00'
        
        ws.cell(row=row_idx, column=4, value=fecha_hoy).alignment = Alignment(horizontal="center")

        # Color alterno en filas
        bg_color = "F2F2F2" if idx % 2 == 0 else "FFFFFF"
        for col in range(1, 5):
            c = ws.cell(row=row_idx, column=col)
            c.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            c.border = border_style

    # Auto-ancho de columnas
    column_widths = {'A': 14, 'B': 26, 'C': 20, 'D': 22}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # --- CONFIGURACIÓN DEL GRÁFICO ---
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Cotización de Divisas (VES)"
    chart.y_axis.title = "Monto en VES"
    chart.x_axis.title = "Moneda"
    chart.legend = None

    # Dimensiones para asegurar visibilidad sin cortes
    chart.width = 18
    chart.height = 10

    # Referencia de Datos (Columna C) y Categorías (Columna B)
    max_row = len(tasas) + 3
    data = Reference(ws, min_col=3, min_row=3, max_row=max_row)
    cats = Reference(ws, min_col=2, min_row=4, max_row=max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Configurar etiquetas sobre las barras limpias (solo el número)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showSerName = False
    chart.dataLabels.showCatName = False

    # Insertar el gráfico al lado de la tabla
    ws.add_chart(chart, "F3")

    # Guardado único final
    wb.save(nombre_archivo)
    print(f"Archivo Excel generado con éxito: {nombre_archivo}")

if __name__ == "__main__":
    print("Obteniendo cotizaciones del BCV...")
    datos_tasas = obtener_todas_las_tasas_bcv()
    if datos_tasas:
        exportar_a_excel(datos_tasas)